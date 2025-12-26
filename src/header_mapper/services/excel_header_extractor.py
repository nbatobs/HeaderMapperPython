import openpyxl
from typing import List
from datetime import datetime
from header_mapper.models.sheet_headers import SheetHeaders, ExcelHeaderResult

class ExcelHeaderExtractor:
    """Extracts headers from Excel files, intelligently detecting and merging multiple header rows"""
    
    def extract_headers(self, file_path: str) -> ExcelHeaderResult:
        """Extract all headers from all sheets in an Excel file"""
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        result = ExcelHeaderResult(file_path=file_path)
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_headers = self._extract_sheet_headers(worksheet, sheet_name)
            result.sheets.append(sheet_headers)
        
        workbook.close()
        return result
    
    def _extract_sheet_headers(self, worksheet, sheet_name: str) -> SheetHeaders:
        """Extract headers from a single worksheet, detecting multiple header rows"""
        sheet_headers = SheetHeaders(sheet_name=sheet_name)
        
        if worksheet.max_row == 0 or worksheet.max_column == 0:
            return sheet_headers  # Empty sheet
        
        header_row_count = self._detect_header_row_count(worksheet)
        sheet_headers.header_row_count = header_row_count
        
        for col in range(1, worksheet.max_column + 1):
            merged_header = self._build_merged_header(worksheet, header_row_count, col)
            
            if merged_header.strip():
                sheet_headers.headers.append(merged_header)
        
        return sheet_headers
    
    def _detect_header_row_count(self, worksheet) -> int:
        """
        Detect how many rows contain header information
        Uses heuristics: looks for rows with text values followed by numeric data
        """
        max_header_rows = min(5, worksheet.max_row)
        data_start_row = 1
        
        for row in range(1, max_header_rows + 1):
            has_numeric_data = False
            has_text_data = False
            non_empty_cells = 0
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row, col)
                value = cell.value
                
                if value is not None:
                    non_empty_cells += 1
                    
                    if self._is_numeric_value(value):
                        has_numeric_data = True
                    elif isinstance(value, str) and value.strip():
                        has_text_data = True
            
            # If we find a row with mostly numeric data, headers end before this row
            if has_numeric_data and non_empty_cells > worksheet.max_column * 0.3:
                data_start_row = row
                break
            
            # If row is mostly text, it's likely a header row
            if has_text_data and non_empty_cells > 0:
                data_start_row = row + 1
        
        return max(1, data_start_row - 1)
    
    def _build_merged_header(self, worksheet, header_row_count: int, column: int) -> str:
        """
        Build a merged header string from multiple header rows
        Example: Row1: "%DM", Row2: "Maize DM" -> "%DM Maize DM"
        Handles merged cells: if "%DM" is merged across columns, applies it to all spanned columns
        """
        header_parts = []
        
        for row in range(1, header_row_count + 1):
            cell = worksheet.cell(row, column)
            text = ""
            
            # Check if this cell is part of a merged range
            if self._is_merged_cell(worksheet, row, column):
                merged_value = self._get_merged_cell_value(worksheet, row, column)
                text = merged_value.strip() if merged_value else ""
            else:
                # Regular cell
                if cell.value is not None:
                    text = str(cell.value).strip()
            
            if text:
                header_parts.append(text)
        
        return " ".join(header_parts)
    
    def _is_merged_cell(self, worksheet, row: int, column: int) -> bool:
        """Check if a cell is part of a merged range"""
        for merged_range in worksheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= column <= merged_range.max_col):
                return True
        return False
    
    def _get_merged_cell_value(self, worksheet, row: int, column: int) -> str:
        """Get the value from a merged cell by finding the top-left cell of the merged range"""
        for merged_range in worksheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= column <= merged_range.max_col):
                # Get the value from the top-left cell of the merged range
                top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
                return str(top_left_cell.value) if top_left_cell.value is not None else ""
        return ""
    
    def _is_numeric_value(self, value) -> bool:
        """Check if a value is numeric (int, float, datetime treated as numeric data)"""
        return isinstance(value, (int, float, datetime))
